"""core/gerber_form.py — the measured figures land in the CLIENT'S form.

The request behind this: FCC's F-SAL-01 quotation sheet, with a cell for
every figure Prism measures. The client's form is the format — Prism only
fills the cells whose labels it recognises, never touches a formula, and
leaves everything unmeasured exactly as drawn.
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))), "prism_terminal"))

from core import gerber_form as GF  # noqa: E402

try:
    import openpyxl
    HAVE = True
except Exception:                                   # pragma: no cover
    HAVE = False

FCC = ("/Users/hitarthtrivedi/Documents/PythonProgram/prism-ai-flow/"
       "gerber_test/gerber data/F-SAL-01 Rev-00 QUOTATION FORM.xlsx")

JOB = {
    "answers": {
        "pcb_size_mm": (60.0, 73.0),
        "array_size_mm": (120.0, 73.0),
        "pcbs_per_array": 2,
        "min_track_width_mm": 2.999994,
        "min_track_spacing_mm": 2.043706,
        "min_drill_mm": 0.999998,
        "min_pitch_mm": 5.0,
        "min_smt_pad": "0.30 x 0.60 mm",
        "drill_count": 80,
        "layers": 2,
    },
    "drills": {"tools": [{"dia_mm": 0.999998, "hits": 20},
                         {"dia_mm": 1.3, "hits": 60},
                         {"dia_mm": 9.9, "hits": 0}],   # unused tool
               "total": 80},
}


@unittest.skipUnless(HAVE, "openpyxl not installed")
class TheLabelsFindTheirCells(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.dir = tempfile.mkdtemp()
        cls.tpl = os.path.join(cls.dir, "form.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Board X"
        ws["A2"] = "Board Y"
        ws["C1"] = "Min Line"
        ws["C2"] = "Smallest Hole"
        ws["E1"] = "Pcs / Array"
        ws["E2"] = "Pitch"
        ws["A3"] = "No. of Holes"
        ws["B3"] = "=B20"                 # their arithmetic
        ws["A4"] = "Cu. Wt Finish"        # nothing measured for this
        ws["C4"] = "Min.SMT Length"
        ws["E4"] = "Min.SMT Width"
        ws["A6"] = "HOLE SIZE"
        ws["B6"] = "NO.OF HOLES/ARRAY"
        ws["A7"], ws["B7"] = 8, 0         # pre-printed placeholders
        ws["A8"], ws["B8"] = 10, 0
        ws["A9"], ws["B9"] = 12, 0
        ws["A10"] = "TOTAL DRILL"
        ws["B10"] = "=SUM(B7:B9)"
        wb.save(cls.tpl)
        cls.out = os.path.join(cls.dir, "filled.xlsx")
        cls.result = GF.fill_form(JOB, cls.tpl, cls.out)
        cls.ws = openpyxl.load_workbook(cls.out).active

    def test_each_label_gets_the_value_beside_it_two_decimals(self):
        self.assertEqual(self.ws["B1"].value, 60.0)
        self.assertEqual(self.ws["B2"].value, 73.0)
        self.assertEqual(self.ws["D1"].value, 3.0)
        self.assertEqual(self.ws["D2"].value, 1.0)
        self.assertEqual(self.ws["F1"].value, 2)
        self.assertEqual(self.ws["F2"].value, 5.0)

    def test_smt_length_is_the_longer_side(self):
        self.assertEqual(self.ws["D4"].value, 0.6)
        self.assertEqual(self.ws["F4"].value, 0.3)

    def test_their_formulas_are_never_overwritten(self):
        self.assertEqual(self.ws["B3"].value, "=B20")
        self.assertEqual(self.ws["B10"].value, "=SUM(B7:B9)")

    def test_an_unmeasured_label_is_left_as_drawn(self):
        self.assertIsNone(self.ws["B4"].value)

    def test_the_drill_table_holds_the_real_tools_and_zeroed_leftovers(self):
        # Two used tools written; the unused Ø9.9 (0 hits) never appears;
        # the third placeholder row is zeroed so their SUM stays honest.
        self.assertEqual((self.ws["A7"].value, self.ws["B7"].value), (1.0, 20))
        self.assertEqual((self.ws["A8"].value, self.ws["B8"].value), (1.3, 60))
        self.assertEqual(self.ws["B9"].value, 0)
        self.assertIsNone(self.ws["A9"].value)
        self.assertEqual(self.result["drill_rows"], 2)

    def test_the_template_itself_is_untouched(self):
        ws = openpyxl.load_workbook(self.tpl).active
        self.assertIsNone(ws["B1"].value)
        self.assertEqual(ws["A7"].value, 8)


@unittest.skipUnless(HAVE and os.path.exists(FCC),
                     "the client's real form is not on this machine")
class TheClientsRealForm(unittest.TestCase):
    """Witnessed against FCC's actual F-SAL-01 sheet."""

    @classmethod
    def setUpClass(cls):
        cls.out = os.path.join(tempfile.mkdtemp(), "fcc.xlsx")
        cls.result = GF.fill_form(JOB, FCC, cls.out,
                                  meta={"customer": "Test Co",
                                        "part": "PCB-001"})
        cls.ws = openpyxl.load_workbook(cls.out)["Table 1"]

    def test_the_figures_land_in_fccs_own_cells(self):
        self.assertEqual(self.ws["D8"].value, 60.0)    # Board X
        self.assertEqual(self.ws["D9"].value, 73.0)    # Board Y
        self.assertEqual(self.ws["B18"].value, 3.0)    # Min Line
        self.assertEqual(self.ws["B19"].value, 2.04)   # Min Space
        self.assertEqual(self.ws["B20"].value, 1.0)    # Smallest Hole
        self.assertEqual(self.ws["B8"].value, 2)       # No Layer
        self.assertEqual(self.ws["B7"].value, "Test Co")
        self.assertEqual(self.ws["D7"].value, "PCB-001")

    def test_fccs_totals_still_come_from_fccs_formulas(self):
        self.assertEqual(self.ws["B22"].value, "=B52")
        self.assertEqual(self.ws["B52"].value, "=SUM(B30:B51)")

    def test_the_drill_table_carries_the_measured_tools(self):
        self.assertEqual((self.ws["A30"].value, self.ws["B30"].value),
                         (1.0, 20))
        self.assertEqual((self.ws["A31"].value, self.ws["B31"].value),
                         (1.3, 60))

    def test_the_companys_photos_and_layout_survive_byte_for_byte(self):
        """The client's form is their document — logo, drawing layout,
        print setup and all. openpyxl's save() rebuilt those wrong (the
        logo drawing shrank 12KB → 1KB, printer settings vanished), which
        is why filling is a zip-level patch: every part of the file except
        the one sheet XML must be BYTE-identical to the template."""
        import zipfile
        with zipfile.ZipFile(FCC) as tin, zipfile.ZipFile(self.out) as tout:
            self.assertEqual(set(tin.namelist()), set(tout.namelist()))
            changed = [n for n in tin.namelist()
                       if tin.read(n) != tout.read(n)]
            self.assertEqual(changed, ["xl/worksheets/sheet1.xml"])
            # And even in that one file, the client's own namespace
            # prefixes are untouched — a prefix rename is what makes
            # Excel call the file corrupt.
            sheet = tout.read("xl/worksheets/sheet1.xml").decode()
            self.assertIn('mc:Ignorable="x14ac', sheet)


@unittest.skipUnless(HAVE, "openpyxl not installed")
class TheUnitsFollowTheUsersChoice(unittest.TestCase):
    """Lengths convert to inch or mil at that unit's honest precision;
    counts, layers and text never do."""

    @classmethod
    def setUpClass(cls):
        cls.dir = tempfile.mkdtemp()
        cls.tpl = os.path.join(cls.dir, "form.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Board X"
        ws["A2"] = "Min Line"
        ws["A3"] = "No Layer"
        ws["C1"] = "Pcs / Array"
        ws["A5"] = "HOLE SIZE"
        ws["B5"] = "NO.OF HOLES/ARRAY"
        ws["A9"] = "TOTAL DRILL"
        wb.save(cls.tpl)

    def _fill(self, units):
        out = os.path.join(self.dir, f"{units}.xlsx")
        GF.fill_form(JOB, self.tpl, out, units=units)
        return openpyxl.load_workbook(out).active

    def test_inch_converts_lengths_and_only_lengths(self):
        ws = self._fill("inch")
        self.assertEqual(ws["B1"].value, 2.3622)   # 60 mm
        self.assertEqual(ws["B2"].value, 0.1181)   # 3 mm
        self.assertEqual(ws["B3"].value, 2)        # layers: a count
        self.assertEqual(ws["D1"].value, 2)        # pcs/array: a count
        self.assertEqual(ws["A6"].value, 0.0394)   # Ø1 mm drill
        self.assertEqual(ws["B6"].value, 20)       # hits: a count

    def test_mil_is_the_fabricators_thou(self):
        ws = self._fill("mil")
        self.assertEqual(ws["B1"].value, 2362.2)
        self.assertEqual(ws["A6"].value, 39.37)
        self.assertEqual(ws["A7"].value, 51.18)    # Ø1.3 mm

    def test_mm_stays_exactly_as_before(self):
        ws = self._fill("mm")
        self.assertEqual(ws["B1"].value, 60.0)
        self.assertEqual(ws["A6"].value, 1.0)

    def test_a_made_up_unit_is_refused_with_a_sentence(self):
        with self.assertRaises(GF.FormError):
            GF.fill_form(JOB, self.tpl,
                         os.path.join(self.dir, "x.xlsx"), units="cm")


@unittest.skipUnless(HAVE, "openpyxl not installed")
class EveryExtractableFactLandsAndTheRestStayBlank(unittest.TestCase):
    """Beyond the nine headline figures: annular ring, tool count, which
    sides carry SMT/mask/legend, slots, and the stackup facts the CAM
    tool's own report states. A fact the job does not carry leaves the
    client's cell blank — never guessed."""

    @classmethod
    def setUpClass(cls):
        cls.dir = tempfile.mkdtemp()
        report = os.path.join(cls.dir, "readme.rep")
        with open(report, "w") as f:
            f.write("Material: FR4, board thickness 1.6 mm, copper 1 oz,\n"
                    "surface finish HASL lead free, V-CUT scoring.\n")
        drill = os.path.join(cls.dir, "job.txt")
        with open(drill, "w") as f:
            f.write("M48\nT1C0.018\n%\nT1\nX1Y1\nG85X2Y2\nM30\n")
        cls.job = {
            "answers": {**JOB["answers"],
                        "min_annular_ring_mm": 0.1397,
                        "drill_tools": 10},
            "drills": JOB["drills"],
            "smt": [{"name": "top.gtl", "count": 5}],
            "files": [
                {"name": "top.gtl", "role": "copper_top", "path": ""},
                {"name": "top.gts", "role": "mask_top", "path": ""},
                {"name": "bot.gbs", "role": "mask_bottom", "path": ""},
                {"name": "top.gto", "role": "silk_top", "path": ""},
                {"name": "job.txt", "role": "drill", "path": drill},
                {"name": "readme.rep", "role": "report", "path": report},
            ],
        }
        tpl = os.path.join(cls.dir, "big-form.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        labels = ["Min Annular Ring", "No. of Tools", "Smt Side",
                  "SM Sides", "Legend Sides", "Slots", "Material Type",
                  "Mat. Thickness", "Cu. Wt Finish", "Final Finish",
                  "Scoring", "E.T."]
        for i, label in enumerate(labels, 1):
            ws.cell(row=i, column=1, value=label)
        wb.save(tpl)
        out = os.path.join(cls.dir, "big-filled.xlsx")
        GF.fill_form(cls.job, tpl, out)
        cls.ws = openpyxl.load_workbook(out).active

    def _value(self, label):
        for row in self.ws.iter_rows():
            if row[0].value == label:
                return row[1].value
        raise AssertionError(label)

    def test_the_measured_extras_land(self):
        self.assertEqual(self._value("Min Annular Ring"), 0.14)
        self.assertEqual(self._value("No. of Tools"), 10)
        self.assertEqual(self._value("Smt Side"), "TOP")
        self.assertEqual(self._value("SM Sides"), "BOTH")
        self.assertEqual(self._value("Legend Sides"), "TOP")
        self.assertEqual(self._value("Slots"), "YES")   # the G85 in the file

    def test_the_cam_reports_own_words_fill_the_stackup(self):
        self.assertEqual(self._value("Material Type"), "FR-4")
        self.assertEqual(self._value("Mat. Thickness"), "1.6 mm")
        self.assertEqual(self._value("Cu. Wt Finish"), "1 oz")
        self.assertEqual(self._value("Final Finish"), "HASL")
        self.assertEqual(self._value("Scoring"), "YES")

    def test_what_the_job_does_not_know_stays_blank(self):
        self.assertIsNone(self._value("E.T."))
        # A bare job with no files carries none of the extras.
        bare = GF._job_extras({"answers": {}})
        self.assertIsNone(bare["sm_sides"])
        self.assertIsNone(bare["smt_side"])
        self.assertNotIn("slots", bare)
        self.assertNotIn("material_type", bare)


class TheDialogRemembersTheTemplate(unittest.TestCase):

    def test_the_gui_offers_and_uses_the_clients_format(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        src = open(os.path.join(root, "dialogs", "gerber_dialog.py"),
                   encoding="utf-8").read()
        self.assertIn('cfg.get("gerber_form_template")', src)
        self.assertIn("get_gerber_form", src)
        # The unit choice is offered, remembered, and actually used.
        self.assertIn('cfg.get("gerber_units")', src)
        self.assertIn('addItems(["mm", "inch", "mil"])', src)
        self.assertIn("units=self.cfg.get", src)
        # Filling runs as part of measuring, not as a separate chore.
        measured = src[src.index("def _on_measured("):
                       src.index("def _on_measure_failed(")]
        self.assertIn("_fill_forms", measured)


if __name__ == "__main__":
    unittest.main()
