"""core/stepfile.py — /step, the estimator's first hour done offline.

Ground truth is the customer's own hand-made drawing sheet for the demo
assembly (step_file_demo/Assem1.STEP, SolidWorks, three sheet-metal
parts): the formed views on its right-hand side give top = 101 x 93 x 71,
side = 92 x 68.5 with 6 mm flanges, and the holes Ø13 / Ø6.2 / Ø9 / Ø5.2.
Those numbers are pinned here as witnessed — if they move, the reading is
wrong, not the drawing.

Synthetic truth-by-construction backs it: a box of known size with one
known hole, exported to STEP and read back.
"""
from __future__ import annotations

import os
import re
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))), "prism_terminal"))

from core import stepfile as SF  # noqa: E402

HAVE = SF.available()[0]
REAL = ("/Users/hitarthtrivedi/Documents/PythonProgram/prism-ai-flow/"
        "step_file_demo/Assem1.STEP")


def _box_with_hole(path: str):
    """A 60 x 40 x 8 mm block with one Ø5 through hole — every figure
    known by construction."""
    import cadquery as cq
    part = (cq.Workplane("XY").box(60, 40, 8)
            .faces(">Z").workplane().hole(5))
    cq.exporters.export(part, path)


@unittest.skipUnless(HAVE, "cadquery not installed")
class ABoxOfKnownSize(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.path = os.path.join(tempfile.mkdtemp(), "block.step")
        _box_with_hole(cls.path)
        cls.report = SF.analyse(cls.path, mode="plastic")

    def test_the_size_is_the_boxs(self):
        self.assertEqual(len(self.report["parts"]), 1)
        self.assertEqual(self.report["parts"][0]["size_mm"], (60.0, 40.0, 8.0))
        self.assertEqual(self.report["overall_mm"], (60.0, 40.0, 8.0))

    def test_the_hole_is_found_once_at_its_diameter(self):
        holes = self.report["parts"][0]["holes"]
        self.assertEqual(holes, [{"dia_mm": 5.0, "count": 1}])

    def test_volume_is_the_boxs_minus_the_hole(self):
        expected = (60 * 40 * 8 - 3.14159 * 2.5 ** 2 * 8) / 1000
        self.assertAlmostEqual(self.report["parts"][0]["volume_cm3"],
                               expected, delta=0.05)

    def test_plastic_mode_prices_the_shot_not_the_sheet(self):
        text = SF.report_text(self.report)
        self.assertIn("plastic moulding", text)
        self.assertIn("wall≈", text)
        self.assertIn("ABS", text)

    def test_every_figure_is_two_decimals(self):
        for number in re.findall(r"\d+\.(\d+)", SF.report_text(self.report)):
            self.assertLessEqual(len(number), 2)


@unittest.skipUnless(HAVE, "cadquery not installed")
class TheBriefForTheImageAgent(unittest.TestCase):
    """/step-auto's prompt: exact measured figures in, the model kept out."""

    @classmethod
    def setUpClass(cls):
        cls.path = os.path.join(tempfile.mkdtemp(), "block.step")
        _box_with_hole(cls.path)
        cls.brief = SF.auto_brief(SF.analyse(cls.path, mode="plastic"))

    def test_the_measured_figures_are_in_it_verbatim(self):
        self.assertIn("60.00 x 40.00 x 8.00 mm", self.brief)
        self.assertIn("Ø5 x 1", self.brief)

    def test_it_says_the_model_is_not_shared(self):
        self.assertIn("MEASURED OFFLINE", self.brief)
        self.assertIn("is not shared", self.brief)

    def test_it_forbids_invented_numbers(self):
        self.assertIn("EXACTLY the numbers above", self.brief)
        self.assertIn("do not round", self.brief)

    def test_it_forbids_fake_flat_patterns_and_tolerances(self):
        """ChatGPT's first real sheet labelled a decorative view 'FLAT
        PATTERN' and invented a ±0.05 tolerance note — both are the kind of
        plausible extra a fabricator would act on."""
        self.assertIn("Never label any view 'flat pattern'", self.brief)
        self.assertIn("Do not state any tolerance", self.brief)

    def test_every_figure_is_two_decimals(self):
        for number in re.findall(r"\d+\.(\d+)", self.brief):
            self.assertLessEqual(len(number), 2)


@unittest.skipUnless(HAVE and os.path.exists(REAL),
                     "cadquery or the demo assembly not on this machine")
class TheCustomersOwnEnclosure(unittest.TestCase):
    """Witnessed against the fab's hand-made drawing sheet."""

    @classmethod
    def setUpClass(cls):
        cls.report = SF.analyse(REAL, mode="metal")
        cls.by_name = {p["name"]: p for p in cls.report["parts"]}

    def test_the_three_parts_keep_their_names(self):
        self.assertEqual(set(self.by_name), {"top", "bottom", "side"})

    def test_the_top_cover_matches_the_drawing(self):
        self.assertEqual(self.by_name["top"]["size_mm"], (101.0, 93.0, 71.0))

    def test_the_side_panel_matches_the_drawing(self):
        self.assertEqual(self.by_name["side"]["size_mm"], (92.0, 68.5, 6.0))

    def test_the_drawings_named_holes_are_all_found(self):
        top = {h["dia_mm"] for h in self.by_name["top"]["holes"]}
        bottom = {h["dia_mm"] for h in self.by_name["bottom"]["holes"]}
        self.assertIn(13.0, top)
        self.assertIn(6.2, top)
        self.assertIn(9.0, bottom)
        self.assertIn(5.2, bottom)

    def test_the_sheet_reads_as_one_millimetre(self):
        for part in self.report["parts"]:
            self.assertGreater(part["thickness_mm"], 0.8)
            self.assertLess(part["thickness_mm"], 1.05)

    def test_the_report_says_formed_not_flat(self):
        self.assertIn("FORMED", SF.report_text(self.report))

    def test_the_excel_sheet_carries_every_part(self):
        import openpyxl
        out = os.path.join(tempfile.mkdtemp(), "dimensions.xlsx")
        SF.write_xlsx(self.report, out)
        wb = openpyxl.load_workbook(out)
        body = "\n".join(str(c.value) for row in wb["Parts"].iter_rows()
                         for c in row if c.value is not None)
        for name in ("top", "bottom", "side"):
            self.assertIn(name, body)
        self.assertIn("101.0", body)
        self.assertEqual(wb["Holes"].max_row - 1,
                         sum(len(p["holes"]) for p in self.report["parts"]))

    def test_the_drawing_sheet_shows_every_part(self):
        out = tempfile.mkdtemp()
        drawn = SF.render_sheet(self.report, out)
        html = open(drawn["html"], encoding="utf-8").read()
        for name in ("top", "bottom", "side"):
            self.assertIn(name, html)
            svg = os.path.join(out, f"{name}.svg")
            self.assertTrue(os.path.exists(svg), svg)
            self.assertGreater(os.path.getsize(svg), 5000, svg)
        self.assertIn("101.00 x 93.00 x 71.00", html)


class ThePlanIsValidatedNotTrusted(unittest.TestCase):
    """/step-ask: whatever the agent answers, only the two executable ops
    survive, and only with sane numbers."""

    def test_a_fenced_json_plan_parses(self):
        plan, _ = SF.parse_plan([
            'Here you go:\n```json\n{"changes": [{"op": "enlarge_hole", '
            '"part": "top", "dia_mm": 5, "new_dia_mm": 6.5, "why": "M6"}], '
            '"advice": ["add draft"]}\n```'])
        self.assertEqual(plan["changes"], [
            {"op": "enlarge_hole", "part": "top", "dia_mm": 5.0,
             "new_dia_mm": 6.5, "why": "M6"}])
        self.assertEqual(plan["advice"], ["add draft"])

    def test_a_shrink_or_unknown_op_is_dropped(self):
        plan, _ = SF.parse_plan([
            '{"changes": [{"op": "enlarge_hole", "part": "x", "dia_mm": 6, '
            '"new_dia_mm": 5}, {"op": "delete_part", "part": "x"}, '
            '{"op": "scale", "part": "x", "factor": 99}], "advice": []}'])
        self.assertEqual(plan["changes"], [])

    def test_garbage_returns_none_with_a_sentence(self):
        plan, why = SF.parse_plan(["no json here", ""])
        self.assertIsNone(plan)
        self.assertIn("no JSON plan", why)

    def test_the_newest_capture_wins(self):
        older = '{"changes": [], "advice": ["old"]}'
        newer = '{"changes": [], "advice": ["new"]}'
        plan, _ = SF.parse_plan([older, newer])
        self.assertEqual(plan["advice"], ["new"])


@unittest.skipUnless(HAVE, "cadquery not installed")
class ThePromptsCarryNumbersNotTheModel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.path = os.path.join(tempfile.mkdtemp(), "block.step")
        _box_with_hole(cls.path)
        cls.report = SF.analyse(cls.path, mode="plastic")

    def test_groq_hears_the_figures_and_the_confidentiality(self):
        p = SF.ask_prompt(self.report, "make it lighter")
        self.assertIn("60.00 x 40.00 x 8.00", p)
        self.assertIn("make it lighter", p)
        self.assertIn("cannot be shown to you", p)

    def test_the_planner_gets_the_schema_and_the_honest_out(self):
        p = SF.plan_prompt(self.report, "q", "advice text")
        self.assertIn('"enlarge_hole"', p)
        self.assertIn('"scale"', p)
        self.assertIn("never shrunk", p)
        self.assertIn("empty changes list", p)


@unittest.skipUnless(HAVE, "cadquery not installed")
class ChangesLandOnACopy(unittest.TestCase):
    """apply_plan edits real geometry and the re-measure proves it."""

    @classmethod
    def setUpClass(cls):
        cls.dir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.dir, "block.step")
        _box_with_hole(cls.path)

    def test_a_hole_is_enlarged_and_nothing_else_moves(self):
        out = os.path.join(self.dir, "bigger.step")
        done = SF.apply_plan(self.path, {"changes": [
            {"op": "enlarge_hole", "part": "all", "dia_mm": 5.0,
             "new_dia_mm": 6.0, "why": ""}], "advice": []}, out)
        self.assertTrue(any("1 hole(s) enlarged" in l for l in done["log"]))
        part = SF.analyse(out, mode="plastic")["parts"][0]
        self.assertEqual(part["size_mm"], (60.0, 40.0, 8.0))
        self.assertEqual(part["holes"], [{"dia_mm": 6.0, "count": 1}])

    def test_a_scaled_part_keeps_its_holes_as_holes(self):
        out = os.path.join(self.dir, "scaled.step")
        SF.apply_plan(self.path, {"changes": [
            {"op": "scale", "part": "all", "factor": 1.5, "why": ""}],
            "advice": []}, out)
        part = SF.analyse(out, mode="plastic")["parts"][0]
        self.assertEqual(part["size_mm"], (90.0, 60.0, 12.0))
        self.assertEqual(part["holes"], [{"dia_mm": 7.5, "count": 1}])

    def test_the_original_file_is_untouched(self):
        before = open(self.path, "rb").read()
        SF.apply_plan(self.path, {"changes": [
            {"op": "scale", "part": "all", "factor": 2, "why": ""}],
            "advice": []}, os.path.join(self.dir, "x.step"))
        self.assertEqual(open(self.path, "rb").read(), before)

    def test_a_plan_that_lands_nowhere_refuses_to_write(self):
        out = os.path.join(self.dir, "never.step")
        with self.assertRaises(SF.StepError):
            SF.apply_plan(self.path, {"changes": [
                {"op": "enlarge_hole", "part": "no-such-part",
                 "dia_mm": 5.0, "new_dia_mm": 6.0, "why": ""}],
                "advice": []}, out)
        self.assertFalse(os.path.exists(out))


@unittest.skipUnless(HAVE and os.path.exists(REAL),
                     "cadquery or the demo assembly not on this machine")
class ChangesOnTheCustomersEnclosure(unittest.TestCase):
    """Witnessed on the real assembly: one hole grows, names and every
    other figure hold still."""

    def test_only_the_named_hole_on_the_named_part_changes(self):
        out = os.path.join(tempfile.mkdtemp(), "m.step")
        SF.apply_plan(REAL, {"changes": [
            {"op": "enlarge_hole", "part": "top", "dia_mm": 6.2,
             "new_dia_mm": 8.0, "why": ""}], "advice": []}, out)
        by = {p["name"]: p for p in SF.analyse(out, "metal")["parts"]}
        self.assertEqual(set(by), {"top", "bottom", "side"})
        top = {h["dia_mm"] for h in by["top"]["holes"]}
        self.assertIn(8.0, top)
        self.assertNotIn(6.2, top)
        self.assertEqual(by["top"]["size_mm"], (101.0, 93.0, 71.0))
        self.assertEqual(by["side"]["size_mm"], (92.0, 68.5, 6.0))
        self.assertIn(5.2, {h["dia_mm"] for h in by["bottom"]["holes"]})


class TheTerminalDoor(unittest.TestCase):

    def test_step_and_s_reach_the_command(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        src = open(os.path.join(root, "prism_terminal", "prism.py"),
                   encoding="utf-8").read()
        self.assertIn("def cmd_step(", src)
        self.assertIn('line.startswith("/step")', src)
        self.assertIn('line == "/s" or line.startswith("/s ")', src)
        self.assertIn("no AI sees the STEP file", src)

    def test_the_mode_words_are_metal_and_plastic(self):
        self.assertEqual(SF.MODES, ("metal", "plastic"))

    def test_step_auto_is_dispatched_before_step(self):
        """startswith("/step") would swallow "/step-auto" — the auto branch
        must be checked first or the command silently runs plain /step."""
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        src = open(os.path.join(root, "prism_terminal", "prism.py"),
                   encoding="utf-8").read()
        self.assertIn("def cmd_step_auto(", src)
        self.assertLess(src.index('line.startswith("/step-auto")'),
                        src.index('line.startswith("/step") or'))

    def test_step_auto_never_uploads_the_model(self):
        """The only attachment /step-auto builds is Prism's OWN render; the
        customer's STEP file must never be in the file list."""
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        src = open(os.path.join(root, "prism_terminal", "prism.py"),
                   encoding="utf-8").read()
        body = src[src.index("def cmd_step_auto("):
                   src.index("def cmd_step_ask(")]
        self.assertIn('F.attach(drawn["png"])', body)
        self.assertNotIn("F.attach(target", body)
        self.assertIn("STEP file stays here", body)

    def test_step_ask_is_dispatched_and_never_uploads_the_model(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        src = open(os.path.join(root, "prism_terminal", "prism.py"),
                   encoding="utf-8").read()
        self.assertLess(src.index('line.startswith("/step-ask")'),
                        src.index('line.startswith("/step") or'))
        body = src[src.index("def cmd_step_ask("):
                   src.index("def cmd_gerber(")]
        self.assertNotIn("F.attach(target", body)
        self.assertIn('F.attach(drawn["png"])', body)
        self.assertIn("groq_chat", body)
        self.assertIn('"modified.step"', body)


if __name__ == "__main__":
    unittest.main()
